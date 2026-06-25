#!/usr/bin/env python3
"""
build_quiz_data.py
==================================================================
Turn a CSV or Excel question bank into the quiz-data.js file the
Bible Quiz app reads. Run this instead of editing quiz-data.js by hand.

USAGE
  python build_quiz_data.py questions.csv
  python build_quiz_data.py questions.xlsx --book "Acts" --out quiz-data.js
  python build_quiz_data.py questions.csv --church "Okota Baptist Church" --version NIV

INPUT FORMAT  (one row per question)
  Required columns (case-insensitive, any order):
    category   - short code or label, e.g. general / tof / fill / golden / recitation
    question   - the question text   (use a literal \n for a line break)
    answer     - the answer text

  Optional columns:
    sn         - ignored for ordering unless --sort-by-sn is passed
    order      - integer; if present, questions sort by it within a category

  Category codes are mapped to friendly labels and ordered on screen via
  CATEGORY_MAP below. Add to it if you invent new categories. Any code not
  in the map is title-cased and appended after the known ones.

  A category whose code/label contains "recit" (or is listed in
  INSTRUCTION_CATEGORIES) is treated as an INSTRUCTION round: it renders as a
  single card that shows its text immediately (no hidden answer, no number grid).
  For those rows, put the round title in `question` and the rules in `answer`.
==================================================================
"""

import argparse
import json
import sys
import os
import re
import random
from collections import OrderedDict

# ------------------------------------------------------------------
# Category code -> (label, on-screen order). Edit freely.
# Keys are matched case-insensitively against the CSV's category cell.
# ------------------------------------------------------------------
CATEGORY_MAP = OrderedDict([
    ("general",        ("General Questions", 1)),
    ("recitation",     ("Recitation",        2)),
    ("recite",         ("Recitation",        2)),
    ("fill",           ("Fill in the Gap",   3)),
    ("fill-in-the-gap",("Fill in the Gap",   3)),
    ("fitg",           ("Fill in the Gap",   3)),
    ("tof",            ("True or False",     4)),
    ("true-or-false",  ("True or False",     4)),
    ("tf",             ("True or False",     4)),
    ("golden-rush",    ("Spelling",          5)),
    ("golden",         ("Spelling",          5)),
    ("spell",          ("Spelling",          5)),
    ("spelling",       ("Spelling",          5)),
])


# NOTE: Whether a category is an "instruction round" (a single always-visible
# card, e.g. a rules slide) is NOT guessed from its name. A category named
# "Recitation" is a NORMAL question grid by default. To make a category an
# instruction card, add an `instruction` column to the CSV and put yes/true/1
# in that column for the relevant row(s). See is_truthy() and build().


def norm(s):
    """Trim and collapse runs of spaces/tabs, but PRESERVE newlines.
    Also accept a literal backslash-n in a cell as an explicit line break."""
    s = str(s)
    # convert a literal two-char sequence \n (common when authoring in Excel) to a real newline
    s = s.replace("\\n", "\n")
    # collapse spaces/tabs (not newlines)
    s = re.sub(r"[ \t]+", " ", s)
    # trim trailing/leading spaces on each line, and overall
    s = "\n".join(line.strip() for line in s.split("\n"))
    return s.strip()


def norm_key(s):
    """Aggressive normalize for header/category keys (newlines irrelevant here)."""
    return re.sub(r"\s+", " ", str(s)).strip()


def is_truthy(v):
    """Interpret a cell as a yes/no flag."""
    return norm_key(v).lower() in {"1", "y", "yes", "true", "t", "instruction", "card"}


def cat_key(raw):
    """Normalise a category cell to a lookup key."""
    k = norm_key(raw).lower().replace(" ", "-")
    return k


def resolve_category(raw):
    """Return (id, label, order) for a raw category cell."""
    k = cat_key(raw)
    if k in CATEGORY_MAP:
        label, order = CATEGORY_MAP[k]
        cid = re.sub(r"[^a-z0-9]+", "-", label.lower()).strip("-")
        return cid, label, order
    # Unknown category: derive a label, place it after known ones
    label = norm_key(raw).title()
    cid = re.sub(r"[^a-z0-9]+", "-", label.lower()).strip("-") or "category"
    return cid, label, 99


def read_rows(path):
    """Read CSV or Excel into a list of dict rows with lowercased keys."""
    ext = os.path.splitext(path)[1].lower()
    if ext in (".xlsx", ".xlsm", ".xls"):
        try:
            from openpyxl import load_workbook
        except ImportError:
            sys.exit("ERROR: reading Excel needs openpyxl. Install with:\n"
                     "  pip install openpyxl --break-system-packages\n"
                     "Or save your file as .csv and re-run.")
        if ext == ".xls":
            sys.exit("ERROR: legacy .xls not supported. Re-save as .xlsx or .csv.")
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        try:
            header = next(rows_iter)
        except StopIteration:
            return []
        keys = [norm_key(h).lower() if h is not None else "" for h in header]
        out = []
        for r in rows_iter:
            if r is None:
                continue
            d = {keys[i]: ("" if v is None else v) for i, v in enumerate(r) if i < len(keys)}
            if any(norm(v) for v in d.values()):
                out.append(d)
        return out
    else:
        import csv
        with open(path, newline="", encoding="utf-8-sig") as f:
            reader = csv.DictReader(f)
            reader.fieldnames = [norm_key(h).lower() for h in (reader.fieldnames or [])]
            out = []
            for row in reader:
                d = {(k or ""): (v if v is not None else "") for k, v in row.items()}
                if any(norm(v) for v in d.values()):
                    out.append(d)
            return out


def pick(row, *names):
    """First non-empty value among the given column names."""
    for n in names:
        if n in row and norm(row[n]):
            return norm(row[n])
    return ""


def js_string(s):
    """JSON-encode a string for embedding in JS (keeps unicode, escapes quotes)."""
    # json.dumps gives us a valid double-quoted JS string literal.
    return json.dumps(s, ensure_ascii=False)


def build(rows, sort_by_sn=False, shuffle=False, seed=None):
    """Group rows into ordered categories.
    The CSV stays in its tidy order; shuffling only affects the OUTPUT order of
    questions within each category (the on-screen tile order), so the source
    file remains readable. A seed makes the shuffle reproducible."""
    buckets = OrderedDict()   # cid -> dict(label, order, is_instruction, items)
    errors = []

    for i, row in enumerate(rows, start=2):  # row 1 = header
        cat = pick(row, "category", "cat", "type")
        q = pick(row, "question", "q", "prompt")
        a = pick(row, "answer", "a", "ans")
        if not cat:
            errors.append(f"row {i}: missing category -> skipped")
            continue
        if not q:
            errors.append(f"row {i}: missing question -> skipped")
            continue
        if not a:
            # allow empty answer only for instruction rows handled below
            a = ""
        cid, label, order = resolve_category(cat)
        # instruction flag comes from an explicit column, not the category name
        row_is_instr = is_truthy(pick(row, "instruction", "is_instruction", "card", "rules"))
        b = buckets.get(cid)
        if b is None:
            b = {"id": cid, "label": label, "order": order,
                 "is_instruction": row_is_instr, "items": []}
            buckets[cid] = b
        elif row_is_instr:
            b["is_instruction"] = True   # any opted-in row makes the category an instruction card
        # sort key
        sk = None
        if sort_by_sn:
            sk = pick(row, "sn", "no", "number")
        if not sk:
            sk = pick(row, "order")
        try:
            sk = int(sk) if sk else None
        except ValueError:
            sk = None
        b["items"].append({"q": q, "a": a, "_sk": sk, "_seq": len(b["items"])})

    # finalize ordering
    cats = sorted(buckets.values(), key=lambda c: (c["order"], c["label"]))
    rng = random.Random(seed) if shuffle else None
    for ci, c in enumerate(cats):
        # sort items by sort-key if any present, else keep input order
        if any(it["_sk"] is not None for it in c["items"]):
            c["items"].sort(key=lambda it: (it["_sk"] is None, it["_sk"] if it["_sk"] is not None else 0, it["_seq"]))
        for it in c["items"]:
            it.pop("_sk", None); it.pop("_seq", None)
        # shuffle the on-screen order within this category (source CSV stays tidy).
        # Derive a distinct sub-seed per category so categories don't all permute
        # identically, while the whole run stays reproducible from one --seed.
        if rng is not None:
            sub = random.Random((seed if seed is not None else 0) * 1000 + ci)
            sub.shuffle(c["items"])
        # instruction categories: warn if more than one row (a rules card expects one)
        if c["is_instruction"] and len(c["items"]) > 1:
            errors.append(f"note: '{c['label']}' is marked as an instruction round but has "
                          f"{len(c['items'])} rows; an instruction round is usually a single "
                          f"card (all rows kept, each renders as its own card).")
    return cats, errors


def render_js(cats, meta):
    lines = []
    lines.append("/* ============================================================")
    lines.append("   BIBLE QUIZ - DATA FILE  (AUTO-GENERATED - do not edit by hand)")
    lines.append("   ------------------------------------------------------------")
    lines.append("   Generated by build_quiz_data.py from your question bank.")
    lines.append("   To change questions, edit the CSV/Excel and re-run the script.")
    lines.append("   ============================================================ */")
    lines.append("")
    lines.append("const QUIZ_META = {")
    lines.append(f"  book:    {js_string(meta['book'])},")
    lines.append(f"  version: {js_string(meta['version'])},")
    lines.append(f"  church:  {js_string(meta['church'])},")
    lines.append(f"  title:   {js_string(meta['title'])},")
    lines.append(f"  hostPassword: {js_string(meta['host_password'])},  // password to open Host view")
    lines.append("};")
    lines.append("")
    lines.append("const CATEGORIES = [")
    for c in cats:
        lines.append("  {")
        lines.append(f"    id: {js_string(c['id'])},")
        lines.append(f"    label: {js_string(c['label'])},")
        if c["is_instruction"]:
            lines.append("    isInstruction: true,")
        lines.append("    items: [")
        for it in c["items"]:
            lines.append(f"      {{ q: {js_string(it['q'])}, a: {js_string(it['a'])} }},")
        lines.append("    ],")
        lines.append("  },")
    lines.append("];")
    lines.append("")
    return "\n".join(lines)


def main():
    ap = argparse.ArgumentParser(description="Build quiz-data.js from CSV/Excel.")
    ap.add_argument("input", help="Path to .csv or .xlsx question bank")
    ap.add_argument("--out", default="quiz-data.js", help="Output JS file (default: quiz-data.js)")
    ap.add_argument("--book", default=None, help="Book name (e.g. Hebrews)")
    ap.add_argument("--version", default="NIV", help="Bible version (default: NIV)")
    ap.add_argument("--church", default="Okota Baptist Church", help="Church name")
    ap.add_argument("--title", default="Bible Quiz", help="Title shown in the header")
    ap.add_argument("--password", default="okota",
                    help="Password to open Host view (default: okota)")
    ap.add_argument("--sort-by-sn", action="store_true",
                    help="Order questions within a category by the 'sn' column")
    ap.add_argument("--shuffle", action="store_true",
                    help="Randomise the on-screen order of questions within each "
                         "category (the CSV file itself stays in tidy order)")
    ap.add_argument("--seed", type=int, default=None,
                    help="Random seed for --shuffle, so the same shuffle can be "
                         "reproduced (e.g. --shuffle --seed 42)")
    args = ap.parse_args()

    if not os.path.exists(args.input):
        sys.exit(f"ERROR: file not found: {args.input}")

    rows = read_rows(args.input)
    if not rows:
        sys.exit("ERROR: no data rows found. Is the header row present?")

    cats, errors = build(rows, sort_by_sn=args.sort_by_sn,
                         shuffle=args.shuffle, seed=args.seed)
    if not cats:
        sys.exit("ERROR: no usable questions found. Check your 'category', "
                 "'question', and 'answer' columns.")

    # infer book name if not given
    book = args.book
    if not book:
        base = os.path.splitext(os.path.basename(args.input))[0]
        book = re.sub(r"[_\-]+", " ", base).replace("quiz", "").strip().title() or "Quiz"

    meta = {"book": book, "version": args.version,
            "church": args.church, "title": args.title,
            "host_password": args.password}

    js = render_js(cats, meta)
    with open(args.out, "w", encoding="utf-8") as f:
        f.write(js)

    # report
    total = sum(len(c["items"]) for c in cats)
    print(f"\u2713 Wrote {args.out}")
    print(f"  Book: {book}  |  Version: {args.version}  |  Church: {args.church}")
    if args.shuffle:
        print(f"  Shuffle: ON" + (f" (seed {args.seed})" if args.seed is not None else " (random seed)"))
    print(f"  Categories: {len(cats)}   Questions: {total}")
    for c in cats:
        tag = "  (instruction)" if c["is_instruction"] else ""
        print(f"    - {c['label']:<20} {len(c['items']):>3}{tag}")
    if errors:
        print("\n  Notes / warnings:")
        for e in errors:
            print(f"    \u2022 {e}")


if __name__ == "__main__":
    main()
